import os.path, shutil, pytest
from io import BytesIO
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from pandas import read_csv


@pytest.fixture
def folders():
    if not os.path.exists('resources'):
        os.mkdir('resources')
    if not os.path.exists('tmp'):
        os.mkdir('tmp')
#
@pytest.fixture
def archiving():
    if not os.path.exists('resources/test_archive.zip'):
        shutil.make_archive('test_archive', 'zip', 'tmp')
        shutil.move('test_archive.zip', 'resources/test_archive.zip')


def test_csv(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('look_me.csv')
        data_frame = read_csv(BytesIO(file))
        assert archive.getinfo('look_me.csv').file_size == 57
        assert len(data_frame.axes[0]) == 3
        assert len(data_frame.axes[1]) == 1


def test_pdf(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('pdf.pdf')
        reader = PdfReader(BytesIO(file))
        assert archive.getinfo('pdf.pdf').file_size == 393774
        assert len(reader.pages) == 7
        assert "Google Tag Manager" in reader.pages[0].extract_text()


def test_xlsx(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('look_me.xlsx')
        workbook = load_workbook(BytesIO(file))
        sheet = workbook.active
        assert archive.getinfo('look_me.xlsx').file_size == 9052
        assert len(workbook.sheetnames) == 3
        assert sheet.max_row == 4
        assert sheet.max_column == 3
        assert sheet.cell(row=2, column=2).value == 'My name is Mary'
